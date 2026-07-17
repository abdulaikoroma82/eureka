"""XLSForm exporter.

Purpose
-------
Write the three XLSForm sheets (survey / choices / settings) to a real
``.xlsx`` workbook using openpyxl.  Also exposes an in-memory bytes export
for the Streamlit download button.

Inputs
------
A compiled :class:`~xlsform_studio.models.Questionnaire`.

Outputs
-------
An ``.xlsx`` file on disk (``export``) or a ``bytes`` object
(``export_bytes``).

Example
-------
>>> from xlsform_studio.models import Questionnaire, Question, FormSettings
>>> qn = Questionnaire(settings=FormSettings(form_title="Demo"),
...                    questions=[Question(name="age", xlsform_type="integer", label="Age")])
>>> path = XLSFormExporter().export(qn, "/tmp/demo.xlsx")  # doctest: +SKIP
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..app.config import CHOICES_COLUMNS, SETTINGS_COLUMNS, SURVEY_COLUMNS
from ..models import Questionnaire
from .choices_builder import ChoicesBuilder
from .settings_builder import SettingsBuilder
from .survey_builder import SurveyBuilder

_HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")

#: Leading characters that make Excel / LibreOffice treat a cell as a live
#: formula on open - the CSV/formula-injection surface (``=cmd|...``,
#: ``=WEBSERVICE(...)``, ``=HYPERLINK(...)``, ``@SUM``...). Source
#: questionnaires and AI-authored text are untrusted, so we neutralise them.
#:
#: ``=`` ``@`` and control characters are never valid at the start of *any*
#: XLSForm cell, so they are defused everywhere. ``+`` and ``-`` are defused
#: only in free-text columns: in the expression columns a leading ``-`` is a
#: legitimate unary minus (``-${x}``) or negative default, which must survive.
_ALWAYS_DANGEROUS = ("=", "@", "\t", "\r", "\x00")
_TEXT_DANGEROUS = _ALWAYS_DANGEROUS + ("+", "-")

#: Columns holding arbitrary untrusted prose (plus any ``label::lang`` /
#: ``media::*`` passthrough column, matched by the ``::`` separator).
_FREE_TEXT_COLUMNS = frozenset({"label", "hint", "constraint_message"})


def _defuse(value: object, column: str) -> object:
    """Prefix a text-forcing apostrophe to a value that would otherwise be
    executed as a spreadsheet formula. Excel/LibreOffice strip the apostrophe
    on display but never evaluate the cell. Column-aware so legitimate XLSForm
    expressions (a ``-${x}`` calculation, a ``-1`` default) are preserved."""
    if not isinstance(value, str) or not value:
        return value
    leaders = (_TEXT_DANGEROUS if column in _FREE_TEXT_COLUMNS or "::" in column
               else _ALWAYS_DANGEROUS)
    return "'" + value if value[0] in leaders else value


class XLSFormExporter:
    """Serialise a questionnaire to an XLSForm workbook.

    When a *target* platform is given, the survey sheet is written in that
    platform's column dialect (from ``knowledge/platforms.yaml``) - e.g. for
    SurveyCTO the ``relevant`` header becomes ``relevance`` and
    ``constraint_message`` becomes ``constraint message``, matching
    SurveyCTO's published form template.
    """

    def __init__(self, knowledge=None) -> None:
        self.survey_builder = SurveyBuilder()
        self.choices_builder = ChoicesBuilder()
        self.settings_builder = SettingsBuilder()
        self._kb = knowledge  # lazy: only loaded when a dialect is needed

    # ------------------------------------------------------------------
    def _dialect(self, target: Optional[str]) -> Dict[str, str]:
        if not target:
            return {}
        if self._kb is None:
            from ..engine.knowledge_base import KnowledgeBase
            self._kb = KnowledgeBase.load()
        return dict(self._kb.platform(target).get("dialect", {}) or {})

    def build_workbook(self, questionnaire: Questionnaire,
                       target: Optional[str] = None) -> Workbook:
        wb = Workbook()
        dialect = self._dialect(target)

        survey_rows = self.survey_builder.build(questionnaire)
        choices_rows = self.choices_builder.build(questionnaire)
        settings_rows = self.settings_builder.build(questionnaire)

        # Base columns plus any passthrough columns (translations, media,
        # cascading-select filters) found in the questionnaire.
        survey_cols = SURVEY_COLUMNS + self.survey_builder.extra_columns(questionnaire)
        choices_cols = CHOICES_COLUMNS + self.choices_builder.extra_columns(questionnaire)

        ws_survey = wb.active
        ws_survey.title = "survey"
        self._write_sheet(ws_survey, survey_cols, survey_rows, dialect)

        ws_choices = wb.create_sheet("choices")
        self._write_sheet(ws_choices, choices_cols, choices_rows)

        ws_settings = wb.create_sheet("settings")
        self._write_sheet(ws_settings, SETTINGS_COLUMNS, settings_rows)

        return wb

    def export(self, questionnaire: Questionnaire, path: Union[str, Path],
               target: Optional[str] = None) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = self.build_workbook(questionnaire, target=target)
        wb.save(str(path))
        return path

    def export_bytes(self, questionnaire: Questionnaire,
                     target: Optional[str] = None) -> bytes:
        wb = self.build_workbook(questionnaire, target=target)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    def _write_sheet(self, ws, columns: List[str], rows: List[Dict[str, str]],
                     dialect: Optional[Dict[str, str]] = None) -> None:
        dialect = dialect or {}
        # Header row (renamed to the platform dialect where applicable; the
        # row dicts keep their canonical keys).
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=dialect.get(name, name))
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        # Data rows.
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, col in enumerate(columns, start=1):
                value = row.get(col, "")
                ws.cell(row=r_idx, column=c_idx,
                        value=_defuse(value, col) if value != "" else None)
        self._autosize(ws, columns, rows)
        ws.freeze_panes = "A2"

    @staticmethod
    def _autosize(ws, columns: List[str], rows: List[Dict[str, str]]) -> None:
        # Single pass over the cells (columns x rows), stringifying each value
        # exactly once, instead of one full row scan per column.
        widths = {col: len(col) for col in columns}
        for row in rows:
            for col in columns:
                value = row.get(col, "")
                if value:
                    widths[col] = max(widths[col], len(str(value)))
        for c_idx, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = \
                min(max(widths[col] + 2, 10), 60)
